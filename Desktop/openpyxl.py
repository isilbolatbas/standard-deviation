# -*- coding: utf-8 -*-
"""
Created on Tue Oct 29 15:44:20 2019

@author: isil
"""

from openpyxl import Workbook


wb = Workbook()
ws = wb.active


ws['A1'] = 'AD'
ws['B1'] = 'SOYAD'
ws['C1'] = 'OGRENCI NO'
ws['D1'] = 'TELEFON'

#sonsuz döngü oluşsun diye sonsuza kadar bu sorguları yapabilmen için.
while(1):
    
    secenek = int(input("ogrenci kayit icin 1 silmek icin 2 listele icin 3 cikmak icin 4'e basiniz="))
    if secenek == 1:
    
        ad = str(input("ogrenci ismi="))
        soyad = str(input("ogrenci soyad="))
        ogrenciNo = int(input("ogrenci no="))
        telefon = int(input("telefon="))
        ws.append([ad, soyad, ogrenciNo, telefon])
        wb.save("sample.xlsx")
        
    elif secenek == 2:
        
        sil = int(input("silmek istediginiz ogrencinin numarasini giriniz="))
        
        for i in range(3,ws.max_row+1):
            if ws.cell(row=i, column=3).value == sil:
                ws.delete_rows(i,3)
            else:
                print("bu kayıtta ogrenci bulunamadi")
        
        wb.save("sample.xlsx")
 
    elif secenek == 3:
        
        max_row = ws.max_row
       # max_column = ws.max_column
        
        for i in range(1, max_row+1):
            for j in range(1, 3):
                cell_obj=ws.cell(row=i,column=j)
                print(cell_obj.value,end=' | ')
            print('\n')
        
    else:
        
        break 
   
    
    
